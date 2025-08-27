from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.http import HttpResponse
from ..models import Appearance

def export_apparitions_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Apparitions"
    headers = ["Titre","Playlist","Curateur","Contact","Abonnés","Date d'ajout","Etat","Description","Mise à jour"]
    ws.append(headers)
    for app in Appearance.objects.select_related("track","playlist"):
        row = [
            app.track.name, app.playlist.name, app.playlist.owner_name,
            app.contact, app.playlist.followers, app.added_on.date() if app.added_on else None,
            app.state, app.playlist.description, app.updated_on.date() if app.updated_on else None
        ]
        ws.append(row)
        r = ws.max_row
        if app.playlist.url: ws.cell(row=r, column=2).hyperlink = app.playlist.url
        if app.playlist.owner_url: ws.cell(row=r, column=3).hyperlink = app.playlist.owner_url
    return wb

def export_apparitions_pdf():
    response = HttpResponse(content_type="application/pdf")
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Export des apparitions")
    y -= 30
    p.setFont("Helvetica", 10)
    for app in Appearance.objects.select_related("track","playlist")[:100]:
        text = f"{app.track.name} - {app.playlist.name} ({app.playlist.followers or 'N/A'} abonnés)"
        p.drawString(50, y, text)
        y -= 15
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 50
    p.showPage()
    p.save()
    return response
