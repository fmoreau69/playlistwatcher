import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date

def build_apparitions_preview(file):
    wb = load_workbook(file)
    ws = wb.active
    data = list(ws.values)
    df = pd.DataFrame(data[1:], columns=data[0])

    preview_data = []
    for row_index, row in df.iterrows():
        excel_row = row_index + 2
        playlist_cell = ws.cell(row=excel_row, column=list(df.columns).index("Playlist") + 1)
        curateur_cell = ws.cell(row=excel_row, column=list(df.columns).index("Curateur") + 1)
        playlist_url = playlist_cell.hyperlink.target if playlist_cell.hyperlink else ""
        curateur_url = curateur_cell.hyperlink.target if curateur_cell.hyperlink else ""
        preview_data.append({
            "Titre": row.get("Titre") or "",
            "Playlist": row.get("Playlist") or "",
            "PlaylistURL": playlist_url,
            "Curateur": row.get("Curateur") or "",
            "CurateurURL": curateur_url,
            "Contact": row.get("Contact") or "",
            "Abonnés": (str(row.get("Abonnés")).replace("\u202f","").replace(" ","").strip() if row.get("Abonnés") not in [None,""] else ""),
            "Date d'ajout": clean_preview(row.get("Date d'ajout")),
            "Etat": row.get("Etat") or "",
            "Description": row.get("Description") or "",
            "Mise à jour": clean_preview(row.get("Mise à jour")),
        })
    return preview_data

def build_playlists_preview(file):
    wb = load_workbook(file)
    ws = wb.active
    data = list(ws.values)
    df = pd.DataFrame(data[1:], columns=data[0])

    preview_data = []
    for row_index, row in df.iterrows():
        excel_row = row_index + 2
        url_cell = ws.cell(row=excel_row, column=list(df.columns).index("URL") + 1)
        url = url_cell.hyperlink.target if url_cell.hyperlink else row.get("URL") or ""
        preview_data.append({
            "Nom": row.get("Nom") or "",
            "URL": url,
            "Curateur": row.get("Curateur") or "",
            "Abonnés": row.get("Abonnés") or "",
            "Description": row.get("Description") or ""
        })
    return preview_data

def clean_preview(value):
    """Préparer les valeurs pour affichage dans la preview"""
    if pd.isna(value) or value == "":
        return ""
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)