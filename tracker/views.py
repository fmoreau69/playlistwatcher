import pandas as pd
import threading
import traceback
import json

from django.core.management import call_command
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from openpyxl import load_workbook, Workbook
from datetime import datetime, date
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from spotipy.oauth2 import SpotifyOAuth

from .models import Appearance, Playlist, Artist, Track, TaskStatus, SpotifyCredentials, SpotifyToken
from .forms import TrackForm, ExcelUploadForm, SpotifyCredentialsForm
from tracker.spotify import get_spotify_credentials, client


def dashboard(request):
    rows = Appearance.objects.select_related("track", "playlist").order_by("-updated_on")

    # Actives playlists
    active_playlists = Playlist.objects.count()

    # Nombre de nouvelles playlists découvertes lors du dernier scan
    task_status_scan = TaskStatus.objects.filter(name="scan_playlists").first()
    task_status_discover = TaskStatus.objects.filter(name="discover_playlists").first()

    new_playlists_count = 0
    if task_status_scan and task_status_scan.extra_info:
        try:
            # extra_info format : "created,updated,total"
            new_playlists_count = int(task_status_scan.extra_info.split(",")[0])
        except Exception:
            new_playlists_count = 0

    # Artistes pour le filtre dropdown
    artists = Artist.objects.all()
    main_artist = artists.first() if artists.exists() else None

    # Tracks du main_artist pour initialiser le dropdown
    tracks = Track.objects.filter(artist=main_artist) if main_artist else []

    # Récupération des compteurs depuis extra_json
    scan_progress = task_status_scan.extra_json.get("current", 0) if task_status_scan and task_status_scan.extra_json else 0
    discover_progress = task_status_discover.extra_json.get("current", 0) if task_status_discover and task_status_discover.extra_json else 0

    return render(request, "tracker/dashboard.html", {
        "rows": rows,
        "active_playlists": active_playlists,
        "new_playlists_count": new_playlists_count,
        "artists": artists,
        "main_artist": main_artist,
        "tracks": tracks,  # pour initialiser track-select
        "discover_progress": discover_progress,
        "scan_progress": scan_progress,
        "task_scan": task_status_scan,
        "task_discover": task_status_discover,
    })


# ----- Artiste and track management -----
def artist_track_manage(request):
    artists = Artist.objects.all().order_by("name")
    tracks = Track.objects.select_related("artist").order_by("name")
    return render(request, "tracker/artist_track_form.html", {
        "artists": artists,
        "tracks": tracks,
    })


# ----- Artiste management -----
def artist_list(request):
    artists = Artist.objects.all()
    return render(request, "tracker/artist_list.html", {"artists": artists})

def artist_create(request):
    if request.method == "POST":
        name = request.POST.get("name").strip()
        spotify_id = request.POST.get("spotify_id").strip() or None
        if name:
            artist = Artist.objects.create(name=name, spotify_id=spotify_id)
            messages.success(request, f"Artiste '{artist.name}' ajouté !")
            return redirect("artist_list")
        else:
            messages.error(request, "Le nom de l'artiste est obligatoire.")

    return render(request, "tracker/artist_form.html", {"artist": None})

def artist_update(request, pk):
    artist = get_object_or_404(Artist, pk=pk)

    if request.method == "POST":
        name = request.POST.get("name").strip()
        spotify_id = request.POST.get("spotify_id").strip() or None
        if name:
            artist.name = name
            artist.spotify_id = spotify_id
            artist.save()
            messages.success(request, f"Artiste '{artist.name}' mis à jour !")
            return redirect("artist_list")
        else:
            messages.error(request, "Le nom de l'artiste est obligatoire.")

    return render(request, "tracker/artist_form.html", {"artist": artist})

def artist_delete(request, pk):
    artist = get_object_or_404(Artist, pk=pk)
    name = artist.name
    artist.delete()
    messages.success(request, f"Artiste '{name}' supprimé !")
    return redirect("artist_list")


# ----- Track management -----
def track_list(request):
    tracks = Track.objects.select_related("artist").all()
    return render(request, "tracker/track_list.html", {"tracks": tracks})

def track_create(request):
    if request.method == "POST":
        form = TrackForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(reverse("dashboard"))
    else:
        form = TrackForm()
    return render(request, "tracker/track_form.html", {"form": form})

def track_update(request, pk):
    track = get_object_or_404(Track, pk=pk)

    if request.method == "POST":
        form = TrackForm(request.POST, instance=track)
        if form.is_valid():
            form.save()
            messages.success(request, f"Titre '{track.name}' mis à jour !")
            return redirect("track_list")
        else:
            messages.error(request, "Le formulaire contient des erreurs.")
    else:
        form = TrackForm(instance=track)

    return render(request, "tracker/track_form.html", {"form": form, "track": track})

def track_delete(request, pk):
    track = get_object_or_404(Track, pk=pk)

    if request.method == "POST":
        track.delete()
        messages.success(request, f"Titre '{track.name}' supprimé.")
        return redirect("track_list")

    return render(request, "tracker/track_confirm_delete.html", {"track": track})

def tracks_by_artist(request, artist_id):
    tracks = Track.objects.filter(artist_id=artist_id).values("id", "name")
    return JsonResponse(list(tracks), safe=False)


# ----- Discover playlists -----
def run_discover_playlists_async():
    status, _ = TaskStatus.objects.get_or_create(name="discover_playlists")
    status.status = "running"
    status.stop_requested = False
    status.extra_info = "0"
    status.save()

    try:
        # Appel de la commande discover_playlists
        call_command("discover_playlists")
        status.status = "done"

    except Exception as e:
        status.status = "error"
        status.extra_info = str(e)
        print(f"Erreur globale de la découverte: {e}")
        traceback.print_exc()

    finally:
        status.save()


def run_discover_playlists(request):
    threading.Thread(target=run_discover_playlists_async, daemon=True).start()
    messages.info(request, "Découverte de nouvelles playlists lancée en arrière-plan ⏳")
    return redirect("dashboard")


def stop_discover_playlists(request):
    status = TaskStatus.objects.filter(name="discover_playlists").first()
    if status and status.status == "running":
        status.stop_requested = True
        status.save()
        messages.info(request, "Demande d’arrêt de la découverte envoyée ⏹️")
    else:
        messages.warning(request, "Aucune découverte en cours")
    return redirect("dashboard")


# ----- Scan playlists -----
def run_scan_playlists_async():
    status, _ = TaskStatus.objects.get_or_create(name="scan_playlists")
    status.status = "running"
    status.stop_requested = False
    status.save()

    try:
        # La commande scan_playlists utilise le client interne
        call_command("scan_playlists")
        status.status = "done"

    except Exception as e:
        status.status = "error"
        status.extra_info = str(e)
        print(f"Erreur globale du scan: {e}")
        traceback.print_exc()

    finally:
        status.save()


def run_scan_playlists(request):
    threading.Thread(target=run_scan_playlists_async, daemon=True).start()
    messages.info(request, "Scan des playlists lancé en arrière-plan ⏳")
    return redirect("dashboard")


def stop_scan_playlists(request):
    status = TaskStatus.objects.filter(name="scan_playlists").first()
    if status and status.status == "running":
        status.stop_requested = True
        status.save()
        messages.info(request, "Demande d’arrêt du scan envoyée ⏹️")
    else:
        messages.warning(request, "Aucun scan en cours")
    return redirect("dashboard")


def scan_status(request):
    task_status = TaskStatus.objects.filter(name="scan_playlists").first()
    data = {"status": "idle", "extra_info": "", "current": 0, "total": 0}

    if task_status:
        data["status"] = task_status.status
        data["extra_info"] = task_status.extra_info or ""

        # Si extra_info est du style "X nouvelles, Y mises à jour"
        try:
            parts = (task_status.extra_info or "0 nouvelles, 0 mises à jour").split(",")
            current = int(parts[0].strip().split()[0])  # X
            total = int(parts[1].strip().split()[0])    # Y
            data["current"] = current
            data["total"] = total
        except Exception:
            # fallback si parsing impossible
            data["current"] = 0
            data["total"] = 0

    return JsonResponse(data)


def discover_status(request):
    task_status = TaskStatus.objects.filter(name="discover_playlists").first()
    data = {"status": "idle", "extra_info": "", "current": 0, "total": 0}

    if task_status:
        data["status"] = task_status.status
        data["extra_info"] = task_status.extra_info or ""

        # D'abord on regarde si extra_json est disponible
        if task_status.extra_json and isinstance(task_status.extra_json, dict):
            data["current"] = task_status.extra_json.get("current", 0)
            data["total"] = task_status.extra_json.get("total", 0)
        else:
            # Fallback : tentative de parsing de extra_info
            try:
                # Exemple attendu : "12 nouvelles, 3 maj, 40 explorées"
                parts = (task_status.extra_info or "").split(",")
                # On cherche le nombre avant "explorées"
                explored_part = next((p for p in parts if "explorée" in p), None)
                if explored_part:
                    current = int(explored_part.strip().split()[0])
                    data["current"] = current
                    data["total"] = current  # pas de total dispo
            except Exception:
                pass  # laisse les valeurs par défaut

    return JsonResponse(data)


# Mapping colonnes Excel → champs du modèle Appearance
COLUMN_MAPPING = {
    'Titre': 'title',
    'Playlist': 'playlist',
    'Curateur': 'curator',
    'Contact': 'contact',
    'Abonnés': 'followers',
    'Date d\'ajout': 'added_date',
    'Etat': 'status',
    'Description': 'description',
    'Mise à jour': 'updated_date'
}

def clean_date(value):
    """Nettoyer une date venant de pandas/excel"""
    if pd.isna(value) or value == "":
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def clean_int(value):
    """Nettoie un entier venant d'Excel, retourne None si vide"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).replace("\u202f", "").replace(" ", "").strip()
    return int(s) if s.isdigit() else None

def clean_preview(value):
    """Préparer les valeurs pour affichage dans la preview"""
    if pd.isna(value) or value == "":
        return ""
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)

def import_excel(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]

            # Charger workbook avec openpyxl
            wb = load_workbook(file)
            ws = wb.active

            # Convertir en DataFrame pandas
            data = list(ws.values)
            df = pd.DataFrame(data[1:], columns=data[0])

            # Colonnes obligatoires
            required_columns = [
                "Titre", "Playlist", "Curateur", "Contact", "Abonnés",
                "Date d'ajout", "Etat", "Description", "Mise à jour"
            ]
            for col in required_columns:
                if col not in df.columns:
                    messages.error(request, f"Colonne manquante : {col}")
                    return redirect("import_excel")

            # Construire la liste pour preview
            preview_data = []

            for row_index, row in df.iterrows():
                # Hyperliens
                excel_row = row_index + 2  # 1 = header, +1 pour passer en base 1 Excel
                playlist_cell = ws.cell(row=excel_row, column=list(df.columns).index("Playlist") + 1)
                curateur_cell = ws.cell(row=excel_row, column=list(df.columns).index("Curateur") + 1)
                playlist_url = playlist_cell.hyperlink.target if playlist_cell.hyperlink else ""
                curateur_url = curateur_cell.hyperlink.target if curateur_cell.hyperlink else ""

                preview_data.append({
                    "Titre": row.get("Titre") or "",
                    "Playlist": row.get("Playlist") or "",
                    "PlaylistURL": playlist_url,
                    "Curateur": row.get("Curateur") or "",
                    "CurateurURL": curateur_url,
                    "Contact": row.get("Contact") or "",
                    "Abonnés": (str(row.get("Abonnés")).replace("\u202f", "").replace(" ", "").strip() if row.get("Abonnés") not in [None, ""] else ""),
                    "Date d'ajout": clean_preview(row.get("Date d'ajout")),
                    "Etat": row.get("Etat") or "",
                    "Description": row.get("Description") or "",
                    "Mise à jour": clean_preview(row.get("Mise à jour")),
                })

            # Stocker dans la session
            request.session["import_preview"] = preview_data

            return render(request, "tracker/import_preview.html", {"preview_data": preview_data, "total": len(preview_data)})

    else:
        form = ExcelUploadForm()

    return render(request, "tracker/import_excel.html", {"form": form})

def confirm_import(request):
    data = request.session.get("import_preview")
    if not data:
        messages.error(request, "Aucune donnée à importer.")
        return redirect("import_excel")

    mode = request.POST.get("mode", "complete")  # par défaut compléter
    imported, updated = 0, 0

    for row in data:
        track, _ = Track.objects.get_or_create(
            name=row.get("Titre") or "Inconnu",
            defaults={"spotify_id": f"temp_{(row.get('Titre') or 'unk')}"[:64]}
        )

        playlist, _ = Playlist.objects.get_or_create(
            name=row.get("Playlist") or "Sans nom",
            defaults={
                "spotify_id": f"temp_{(row.get('Playlist') or 'unk')}"[:64],
                "followers": int(row.get("Abonnés")) if row.get("Abonnés") not in ("", None) else None,
                "description": row.get("Description") or "",
                "url": row.get("PlaylistURL") or "",
                "owner_name": row.get("Curateur") or "",
                "owner_url": row.get("CurateurURL") or "",
            }
        )

        added_on = clean_date(row.get("Date d'ajout"))
        updated_on = clean_date(row.get("Mise à jour")) or datetime.today().date()

        appearance, created = Appearance.objects.get_or_create(
            track=track,
            playlist=playlist,
            defaults={
                "contact": row.get("Contact") or "",
                "state": row.get("Etat") or "",
                "added_on": added_on,
                "updated_on": updated_on,
            }
        )

        if created:
            imported += 1
        else:
            if mode == "overwrite":
                appearance.contact = row.get("Contact") or appearance.contact
                appearance.state = row.get("Etat") or appearance.state
                appearance.added_on = added_on or appearance.added_on
                appearance.updated_on = updated_on
                appearance.save()
                updated += 1
            elif mode == "complete":
                changed = False
                if not appearance.contact and row.get("Contact"):
                    appearance.contact = row.get("Contact")
                    changed = True
                if not appearance.state and row.get("Etat"):
                    appearance.state = row.get("Etat")
                    changed = True
                if not appearance.added_on and added_on:
                    appearance.added_on = added_on
                    changed = True
                if changed:
                    appearance.updated_on = updated_on
                    appearance.save()
                    updated += 1

    del request.session["import_preview"]

    messages.success(request, f"{imported} apparitions importées, {updated} mises à jour.")
    return redirect("dashboard")

def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Apparitions"

    headers = ["Titre", "Playlist", "Curateur", "Contact", "Abonnés", "Date d'ajout", "Etat", "Description", "Mise à jour"]
    ws.append(headers)

    for app in Appearance.objects.select_related("track", "playlist"):
        row = [
            app.track.name,
            app.playlist.name,
            app.playlist.owner_name,
            app.contact,
            app.playlist.followers,
            app.added_on,
            app.state,
            app.playlist.description,
            app.updated_on,
        ]
        ws.append(row)
        r = ws.max_row

        if app.playlist.url:
            ws.cell(row=r, column=2).hyperlink = app.playlist.url
        if app.playlist.owner_url:
            ws.cell(row=r, column=3).hyperlink = app.playlist.owner_url

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="export.xlsx"'
    wb.save(response)
    return response

def export_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="export.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Export des apparitions")
    y -= 30

    p.setFont("Helvetica", 10)
    for app in Appearance.objects.select_related("track", "playlist")[:100]:  # limiter pour test
        text = f"{app.track.name} - {app.playlist.name} ({app.playlist.followers or 'N/A'} abonnés)"
        p.drawString(50, y, text)
        y -= 15
        if y < 50:  # nouvelle page
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 50

    p.showPage()
    p.save()
    return response


@login_required
def spotify_login(request):
    creds = get_spotify_credentials()
    sp_oauth = SpotifyOAuth(**creds)

    auth_url = sp_oauth.get_authorize_url()
    return redirect(auth_url)


@login_required  # mieux que csrf_exempt
def spotify_callback(request):
    code = request.GET.get("code")
    error = request.GET.get("error")

    if error:
        return HttpResponseBadRequest(f"Erreur Spotify: {error}")
    if not code:
        return HttpResponseBadRequest("Code d'autorisation manquant.")

    creds = get_spotify_credentials()
    sp_oauth = SpotifyOAuth(**creds)

    try:
        token_info = sp_oauth.get_access_token(code, as_dict=True)
    except Exception as e:
        return HttpResponseBadRequest(f"Erreur lors de l’échange du code: {e}")

    if not token_info:
        return HttpResponseBadRequest("Impossible d'obtenir un token Spotify.")

    expires_at = timezone.now() + timezone.timedelta(seconds=token_info["expires_in"])

    SpotifyToken.objects.update_or_create(
        id=1,
        defaults={
            "access_token": token_info["access_token"],
            "refresh_token": token_info.get("refresh_token", ""),
            "expires_at": expires_at,
        }
    )

    messages.success(request, "Authentification Spotify réussie ✅")
    return redirect("dashboard")  # au lieu de HttpResponse brut


@login_required
def spotify_credentials(request):
    try:
        creds = SpotifyCredentials.objects.get(pk=1)
    except SpotifyCredentials.DoesNotExist:
        creds = SpotifyCredentials()

    if request.method == "POST":
        if "upload_file" in request.FILES:
            f = request.FILES["upload_file"]
            data = json.load(f)

            creds.client_id = data.get("client_id") or data.get("clientId") or ""
            creds.client_secret = data.get("client_secret") or data.get("clientSecret") or ""
            creds.redirect_uri = data.get("redirect_uri") or data.get("redirectUri") or get_spotify_credentials()["redirect_uri"]
            creds.save()

            messages.success(request, "Identifiants importés depuis le fichier.")
            return redirect("spotify_credentials")

        else:
            form = SpotifyCredentialsForm(request.POST, instance=creds)
            if form.is_valid():
                form.save()
                messages.success(request, "Identifiants Spotify enregistrés.")
                return redirect("spotify_credentials")
    else:
        form = SpotifyCredentialsForm(instance=creds)

    return render(request, "tracker/spotify_credentials.html", {"form": form})
